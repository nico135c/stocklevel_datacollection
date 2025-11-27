from arduinoreader import ArduinoReader
from openpyxl import Workbook, load_workbook
import os

def save_part_column(part_no, readings, EXCEL_FILE):
    path = f"data/{EXCEL_FILE}"
    num_readings = len(readings)

    # Create workbook if missing
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # First column: reading indices
        ws.append(["Reading #"] + [])  # row 1 header
        for i in range(1, num_readings + 1):
            ws.append([i])

        wb.save(path)

    # Load workbook
    wb = load_workbook(path)
    ws = wb.active

    # Determine which column this part goes into
    col_index = part_no + 1  # col 1 is Reading #

    # If this is a new column, write header
    header_cell = ws.cell(row=1, column=col_index)
    header_cell.value = f"Part {part_no}"

    # Fill column with readings
    for row in range(2, num_readings + 2):  # Excel row 2 to 201
        ws.cell(row=row, column=col_index).value = readings[row - 2]

    wb.save(path)
    print(f"Saved {num_readings} readings into column for Part {part_no}.\n")


def collect_data(part_amount, reader, EXCEL_FILE):
    for part_no in range(part_amount + 1):
        print(f"\n=== Collecting data for part {part_no}/{part_amount} ===\n")

        readings = reader.get_next(200)

        print("Saving to Excel...\n")
        save_part_column(part_no, readings, EXCEL_FILE)

        print("Done! Insert the next part and press Enter...\n")
        input()
    

if __name__ == "__main__":
    reader = ArduinoReader("/dev/ttyACM0", 115200)
    reader.start()

    collect_data(20, reader, "bottompart_data.xlsx")



